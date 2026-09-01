# reporting/report_writer.py

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill


class ReportWriter:

    def write_report(
        self,
        supplier_result,
        output_file_path
    ):
        workbook = Workbook()

        summary_sheet = workbook.active
        summary_sheet.title = "Summary"

        findings_sheet = workbook.create_sheet(
            "Findings"
        )

        self._write_summary_sheet(
            summary_sheet,
            supplier_result
        )

        self._write_findings_sheet(
            findings_sheet,
            supplier_result
        )

        output_path = Path(output_file_path)

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True
        )

        workbook.save(
            output_file_path
        )

    def _write_summary_sheet(
        self,
        worksheet,
        supplier_result
    ):
        worksheet["A1"] = "Supplier"
        worksheet["B1"] = supplier_result.supplier_name

        actionable_findings = [
            finding
            for finding in supplier_result.findings
            if not getattr(finding, "expected_discrepancy", False)
        ]

        expected_findings = [
            finding
            for finding in supplier_result.findings
            if getattr(finding, "expected_discrepancy", False)
        ]

        worksheet["A3"] = "Total Findings"
        worksheet["B3"] = len(supplier_result.findings)

        worksheet["A4"] = "Actionable (excludes flagged sheets)"
        worksheet["B4"] = len(actionable_findings)

        worksheet["A5"] = (
            "On Sheets Flagged 'Discrepancies Expected'"
        )
        worksheet["B5"] = len(expected_findings)

        row = self._write_severity_table(
            worksheet,
            row=7,
            title="Actionable Findings by Severity",
            findings=actionable_findings,
        )

        row = self._write_severity_table(
            worksheet,
            row=row + 1,
            title="Expected-Sheet Findings by Severity",
            findings=expected_findings,
        )

        row = self._write_category_table(
            worksheet,
            row=row + 1,
            findings=supplier_result.findings,
        )

        coverage = getattr(supplier_result, "benchmark_coverage", None)

        if coverage is not None:
            self._write_coverage_table(
                worksheet,
                row=row + 1,
                coverage=coverage,
            )

        self._bold_row(worksheet, 1)

        worksheet.column_dimensions["A"].width = 40
        worksheet.column_dimensions["B"].width = 20

    def _write_severity_table(
        self,
        worksheet,
        row,
        title,
        findings,
    ):
        worksheet.cell(row=row, column=1).value = title
        self._bold_row(worksheet, row)

        row += 1

        worksheet.cell(row=row, column=1).value = "Severity"
        worksheet.cell(row=row, column=2).value = "Count"
        self._bold_row(worksheet, row)

        severity_counts = {
            "HIGH": 0,
            "MEDIUM": 0,
            "LOW": 0,
            "INFO": 0,
        }

        for finding in findings:
            severity_name = self._severity_name(finding.severity)
            severity_counts.setdefault(severity_name, 0)
            severity_counts[severity_name] += 1

        for severity_name, count in severity_counts.items():
            row += 1
            worksheet.cell(row=row, column=1).value = severity_name
            worksheet.cell(row=row, column=2).value = count

        return row

    def _write_category_table(
        self,
        worksheet,
        row,
        findings,
    ):
        worksheet.cell(row=row, column=1).value = "Findings by Category"
        self._bold_row(worksheet, row)

        row += 1

        worksheet.cell(row=row, column=1).value = "Category"
        worksheet.cell(row=row, column=2).value = "Count"
        self._bold_row(worksheet, row)

        category_counts = {}

        for finding in findings:
            category = getattr(finding, "category", "") or "Uncategorised"
            category_counts[category] = category_counts.get(category, 0) + 1

        for category, count in category_counts.items():
            row += 1
            worksheet.cell(row=row, column=1).value = category
            worksheet.cell(row=row, column=2).value = count

        return row

    def _write_coverage_table(
        self,
        worksheet,
        row,
        coverage,
    ):
        """
        Reports how completely this supplier's submitted fields could
        be checked against the benchmark workbook, so a low match
        rate is visible in the report itself rather than only in the
        mapping viewer - a strong sign the benchmark isn't attached
        correctly for this sheet/field even before any deviations are
        considered.
        """

        worksheet.cell(row=row, column=1).value = "Benchmark Coverage"
        self._bold_row(worksheet, row)

        row += 1

        worksheet.cell(row=row, column=1).value = "Fields Submitted"
        worksheet.cell(row=row, column=2).value = coverage.total_fields

        row += 1

        worksheet.cell(row=row, column=1).value = "Matched to a Benchmark Value"
        worksheet.cell(row=row, column=2).value = coverage.matched_fields

        row += 1

        worksheet.cell(row=row, column=1).value = "Match Rate"
        worksheet.cell(row=row, column=2).value = (
            f"{coverage.match_rate_percent}%"
        )

        if coverage.unmatched_by_sheet:

            row += 2

            worksheet.cell(
                row=row, column=1
            ).value = "Unmatched Fields by Sheet"
            self._bold_row(worksheet, row)

            for sheet_name, count in sorted(
                coverage.unmatched_by_sheet.items(),
                key=lambda item: item[1],
                reverse=True,
            ):
                row += 1
                worksheet.cell(row=row, column=1).value = sheet_name
                worksheet.cell(row=row, column=2).value = count

        if coverage.unmatched_by_field:

            row += 2

            worksheet.cell(
                row=row, column=1
            ).value = "Unmatched Fields by Field Name"
            self._bold_row(worksheet, row)

            for field_name, count in coverage.unmatched_by_field.items():
                row += 1
                worksheet.cell(row=row, column=1).value = field_name
                worksheet.cell(row=row, column=2).value = count

        return row

    def _write_findings_sheet(
        self,
        worksheet,
        supplier_result
    ):
        headers = [
            "Severity",
            "Category",
            "Expected?",
            "Worksheet",
            "Cell",
            "Region",
            "Actual Value",
            "Comparison Basis",
            "Comparator Value",
            "Deviation %",
            "Reason",
            "Suggested Clarification",
        ]

        for column_index, header in enumerate(
            headers,
            start=1
        ):
            cell = worksheet.cell(
                row=1,
                column=column_index
            )

            cell.value = header
            cell.font = Font(
                bold=True
            )

        row_number = 2

        for finding in supplier_result.findings:

            worksheet.cell(
                row=row_number,
                column=1
            ).value = self._severity_name(
                finding.severity
            )

            worksheet.cell(
                row=row_number,
                column=2
            ).value = getattr(
                finding,
                "category",
                ""
            )

            worksheet.cell(
                row=row_number,
                column=3
            ).value = (
                "Yes"
                if getattr(finding, "expected_discrepancy", False)
                else "No"
            )

            worksheet.cell(
                row=row_number,
                column=4
            ).value = getattr(
                finding,
                "worksheet_name",
                ""
            )

            worksheet.cell(
                row=row_number,
                column=5
            ).value = getattr(
                finding,
                "cell_reference",
                ""
            )

            # item_description now stores region/context
            worksheet.cell(
                row=row_number,
                column=6
            ).value = getattr(
                finding,
                "item_description",
                ""
            )

            worksheet.cell(
                row=row_number,
                column=7
            ).value = getattr(
                finding,
                "actual_value",
                ""
            )

            comparator_value = getattr(
                finding,
                "comparator_value",
                None
            )

            worksheet.cell(
                row=row_number,
                column=8
            ).value = getattr(
                finding,
                "comparator_label",
                ""
            ) if comparator_value is not None else ""

            worksheet.cell(
                row=row_number,
                column=9
            ).value = (
                comparator_value
                if comparator_value is not None
                else ""
            )

            worksheet.cell(
                row=row_number,
                column=10
            ).value = getattr(
                finding,
                "deviation_percent",
                None
            )

            worksheet.cell(
                row=row_number,
                column=11
            ).value = getattr(
                finding,
                "reason",
                ""
            )

            worksheet.cell(
                row=row_number,
                column=12
            ).value = getattr(
                finding,
                "suggested_clarification",
                ""
            )

            self._apply_severity_format(
                worksheet.cell(
                    row=row_number,
                    column=1
                )
            )

            row_number += 1

        widths = {
            "A": 15,
            "B": 26,
            "C": 11,
            "D": 25,
            "E": 15,
            "F": 35,
            "G": 20,
            "H": 28,
            "I": 20,
            "J": 15,
            "K": 60,
            "L": 60,
        }

        for column, width in widths.items():
            worksheet.column_dimensions[
                column
            ].width = width

    def _severity_name(
        self,
        severity
    ):
        if hasattr(
            severity,
            "name"
        ):
            return severity.name.upper()

        return str(severity).upper()

    def _apply_severity_format(
        self,
        cell
    ):
        value = str(
            cell.value
        ).upper()

        if value == "HIGH":
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="FF9999"
            )

        elif value == "MEDIUM":
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="FFD699"
            )

        elif value == "LOW":
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="FFF2CC"
            )

        elif value == "INFO":
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="D9EAD3"
            )

    def _bold_row(
        self,
        worksheet,
        row_number
    ):
        for cell in worksheet[row_number]:
            cell.font = Font(
                bold=True
            )