# parsers/workbook_parser.py

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


@dataclass
class CellInfo:
    sheet_name: str
    cell_reference: str
    row_number: int
    column_number: int
    value: Any

    fill_colour: str = ""
    font_name: str = ""
    font_size: float = 0.0

    bold: bool = False
    italic: bool = False
    underline: bool = False

    font_colour: str = ""

    horizontal_alignment: str = ""
    vertical_alignment: str = ""

    number_format: str = ""

    has_formula: bool = False

    is_locked: bool = True
    is_merged: bool = False

    border_left: bool = False
    border_right: bool = False
    border_top: bool = False
    border_bottom: bool = False

    row_height: float | None = None
    column_width: float | None = None


@dataclass
class WorksheetInfo:
    worksheet_name: str

    cells: list[CellInfo] = field(
        default_factory=list
    )

    merged_ranges: list[str] = field(
        default_factory=list
    )

    row_heights: dict[int, float] = field(
        default_factory=dict
    )

    column_widths: dict[str, float] = field(
        default_factory=dict
    )

    def add_cell(
        self,
        cell_info: CellInfo,
    ) -> None:
        self.cells.append(cell_info)

    def get_cell(
        self,
        cell_reference: str,
    ) -> CellInfo | None:

        target = cell_reference.upper()

        for cell in self.cells:
            if cell.cell_reference.upper() == target:
                return cell

        return None


@dataclass
class WorkbookInfo:
    file_name: str
    file_path: str

    worksheets: dict[str, WorksheetInfo] = field(
        default_factory=dict
    )

    def add_worksheet(
        self,
        worksheet_info: WorksheetInfo,
    ) -> None:
        self.worksheets[
            worksheet_info.worksheet_name
        ] = worksheet_info

    def get_worksheet(
        self,
        worksheet_name: str,
    ) -> WorksheetInfo | None:
        return self.worksheets.get(
            worksheet_name
        )


class WorkbookParser:

    def load_workbook(
        self,
        file_path: str,
    ) -> WorkbookInfo:

        path = Path(file_path)

        if not path.exists():
            raise FileNotFoundError(
                f"Workbook not found: {file_path}"
            )

        workbook = load_workbook(
            filename=file_path,
            data_only=False,
            read_only=False,
            keep_vba=(
                path.suffix.lower()
                == ".xlsm"
            ),
        )

        # A formula cell's raw value (data_only=False) is the formula
        # text itself, e.g. "=B2*1.1", not its result - every numeric
        # comparison downstream would then silently fail to parse it
        # as a number and skip the cell entirely, without ever
        # flagging it as blank, unreadable, or worth a clarification.
        # A second, data_only=True load exposes Excel's last-cached
        # calculated result instead (None only if the file was never
        # opened/saved in Excel, in which case the formula text is
        # kept as a fallback - same behaviour as before this fix).
        #
        # That second load re-parses the entire file from scratch
        # (XML, styles, everything) - real money on a large workbook,
        # and most supplier/benchmark workbooks have no formulas in
        # their input cells at all. So it's only opened when a cheap
        # pre-scan of the already-loaded workbook (just checking each
        # cell's value, no CellInfo objects built) finds at least one.
        computed_values_workbook = None

        if self._workbook_has_formulas(workbook):

            computed_values_workbook = load_workbook(
                filename=file_path,
                data_only=True,
                read_only=False,
                keep_vba=False,
            )

        workbook_info = WorkbookInfo(
            file_name=path.name,
            file_path=str(path),
        )

        for worksheet in workbook.worksheets:

            computed_values_worksheet = (
                computed_values_workbook[worksheet.title]
                if computed_values_workbook is not None
                and worksheet.title in computed_values_workbook.sheetnames
                else None
            )

            worksheet_info = (
                self._parse_worksheet(
                    worksheet,
                    computed_values_worksheet,
                )
            )

            workbook_info.add_worksheet(
                worksheet_info
            )

        workbook.close()

        if computed_values_workbook is not None:
            computed_values_workbook.close()

        return workbook_info

    def _workbook_has_formulas(
        self,
        workbook,
    ) -> bool:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:

                    if isinstance(cell, MergedCell):
                        continue

                    if self._is_formula(cell.value):
                        return True

        return False

    def _parse_worksheet(
        self,
        worksheet,
        computed_values_worksheet=None,
    ) -> WorksheetInfo:

        worksheet_info = WorksheetInfo(
            worksheet_name=worksheet.title
        )

        worksheet_info.merged_ranges = [
            str(merged_range)
            for merged_range in worksheet.merged_cells.ranges
        ]

        merged_references = (
            self._get_merged_cell_references(
                worksheet
            )
        )

        for row_number, row_dimension in (
            worksheet.row_dimensions.items()
        ):
            if row_dimension.height is not None:

                worksheet_info.row_heights[
                    int(row_number)
                ] = float(
                    row_dimension.height
                )

        for (
            column_letter,
            column_dimension,
        ) in worksheet.column_dimensions.items():

            if column_dimension.width is not None:

                worksheet_info.column_widths[
                    str(column_letter)
                ] = float(
                    column_dimension.width
                )

        for row in worksheet.iter_rows():

            for cell in row:

                # CRITICAL FIX
                if isinstance(
                    cell,
                    MergedCell,
                ):
                    continue

                formula = self._is_formula(
                    cell.value
                )

                column_letter = (
                    self._column_letter(
                        cell.column
                    )
                )

                cell_value = cell.value

                if formula and computed_values_worksheet is not None:

                    computed_cell = (
                        computed_values_worksheet.cell(
                            row=cell.row,
                            column=cell.column,
                        )
                    )

                    if computed_cell.value is not None:
                        cell_value = computed_cell.value

                cell_info = CellInfo(
                    sheet_name=worksheet.title,
                    cell_reference=cell.coordinate,
                    row_number=cell.row,
                    column_number=cell.column,
                    value=cell_value,

                    fill_colour=self._get_fill_colour(
                        cell
                    ),

                    font_name=str(
                        cell.font.name or ""
                    ),

                    font_size=float(
                        cell.font.sz or 0
                    ),

                    bold=bool(
                        cell.font.bold
                    ),

                    italic=bool(
                        cell.font.italic
                    ),

                    underline=bool(
                        cell.font.underline
                    ),

                    font_colour=self._get_font_colour(
                        cell
                    ),

                    horizontal_alignment=str(
                        cell.alignment.horizontal
                        or ""
                    ),

                    vertical_alignment=str(
                        cell.alignment.vertical
                        or ""
                    ),

                    number_format=str(
                        cell.number_format
                        or ""
                    ),

                    has_formula=formula,

                    is_locked=self._is_locked(
                        cell
                    ),

                    is_merged=(
                        cell.coordinate
                        in merged_references
                    ),

                    border_left=(
                        cell.border.left.style
                        is not None
                    ),

                    border_right=(
                        cell.border.right.style
                        is not None
                    ),

                    border_top=(
                        cell.border.top.style
                        is not None
                    ),

                    border_bottom=(
                        cell.border.bottom.style
                        is not None
                    ),

                    row_height=(
                        worksheet_info.row_heights.get(
                            cell.row
                        )
                    ),

                    column_width=(
                        worksheet_info.column_widths.get(
                            column_letter
                        )
                    ),
                )

                worksheet_info.add_cell(
                    cell_info
                )

        return worksheet_info

    # ==================================================
    # FORMATTING HELPERS
    # ==================================================

    def _get_fill_colour(
        self,
        cell,
    ) -> str:

        try:

            colour = cell.fill.fgColor

            if colour is None:
                return ""

            if (
                colour.type == "rgb"
                and colour.rgb
            ):
                return str(
                    colour.rgb
                ).upper()

            if colour.type == "indexed":
                return (
                    f"INDEXED:{colour.indexed}"
                )

            if colour.type == "theme":
                return (
                    f"THEME:{colour.theme}"
                )

        except Exception:
            pass

        return ""

    def _get_font_colour(
        self,
        cell,
    ) -> str:

        try:

            colour = cell.font.color

            if colour is None:
                return ""

            if (
                colour.type == "rgb"
                and colour.rgb
            ):
                return str(
                    colour.rgb
                ).upper()

            if colour.type == "indexed":
                return (
                    f"INDEXED:{colour.indexed}"
                )

            if colour.type == "theme":
                return (
                    f"THEME:{colour.theme}"
                )

        except Exception:
            pass

        return ""

    # ==================================================
    # FORMULA HELPERS
    # ==================================================

    def _is_formula(
        self,
        value,
    ) -> bool:

        if not isinstance(
            value,
            str,
        ):
            return False

        return value.startswith("=")

    def _has_formula(
        self,
        value,
    ) -> bool:
        """
        Backwards compatibility wrapper.
        """
        return self._is_formula(value)

    # ==================================================
    # LOCKING
    # ==================================================

    def _is_locked(
        self,
        cell,
    ) -> bool:

        try:
            return bool(
                cell.protection.locked
            )

        except Exception:
            return True

    # ==================================================
    # MERGED CELLS
    # ==================================================

    def _get_merged_cell_references(
        self,
        worksheet,
    ) -> set:

        references = set()

        for merged_range in (
            worksheet.merged_cells.ranges
        ):

            for row in worksheet.iter_rows(
                min_row=merged_range.min_row,
                max_row=merged_range.max_row,
                min_col=merged_range.min_col,
                max_col=merged_range.max_col,
            ):
                for cell in row:

                    if isinstance(
                        cell,
                        MergedCell,
                    ):
                        continue

                    references.add(
                        cell.coordinate
                    )

        return references

    # ==================================================
    # COLUMN HELPERS
    # ==================================================

    @staticmethod
    def _column_letter(
        column_number: int,
    ) -> str:

        result = ""

        while column_number > 0:

            column_number, remainder = divmod(
                column_number - 1,
                26,
            )

            result = (
                chr(65 + remainder)
                + result
            )

        return result

    # ==================================================
    # QUERY HELPERS
    # ==================================================

    def get_populated_cells(
        self,
        worksheet_info: WorksheetInfo,
    ) -> list[CellInfo]:

        populated = []

        for cell in worksheet_info.cells:

            if cell.value is None:
                continue

            if (
                  isinstance(cell.value, str)
                  and cell.value.strip() == ""
                ):
                continue

            populated.append(cell)

        return populated