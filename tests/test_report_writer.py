# tests/test_report_writer.py

import openpyxl

from models.pricing_models import (
    BenchmarkCoverage,
    Finding,
    FindingCategory,
    Severity,
    SupplierAnalysisResult,
)
from reporting.report_writer import ReportWriter


def _summary_rows(workbook):
    return [
        row
        for row in workbook["Summary"].iter_rows(values_only=True)
        if any(cell is not None for cell in row)
    ]


def test_write_report_creates_summary_and_findings_sheets(tmp_path):
    findings = [
        Finding(
            severity=Severity.HIGH,
            worksheet_name="Labour",
            cell_reference="B2",
            actual_value="120",
            reason="Value differs from benchmark by 20%",
            category=FindingCategory.BENCHMARK_COMPARISON.value,
            comparator_value="100",
            comparator_label="Benchmark Rate",
            deviation_percent=20.0,
        ),
        Finding(
            severity=Severity.LOW,
            worksheet_name="Extras",
            cell_reference="C4",
            actual_value="",
            reason="Blank value detected.",
            category=FindingCategory.TENDER_RESPONSE_CHECK.value,
            expected_discrepancy=True,
        ),
    ]

    result = SupplierAnalysisResult(supplier_name="Acme Ltd", findings=findings)

    output_path = tmp_path / "acme_report.xlsx"
    ReportWriter().write_report(result, str(output_path))

    assert output_path.exists()

    workbook = openpyxl.load_workbook(output_path)
    assert workbook.sheetnames == ["Summary", "Findings"]

    summary_rows = _summary_rows(workbook)
    assert ("Supplier", "Acme Ltd") in summary_rows
    assert ("Total Findings", 2) in summary_rows
    assert ("Actionable (excludes flagged sheets)", 1) in summary_rows
    assert (
        "On Sheets Flagged 'Discrepancies Expected'", 1
    ) in summary_rows

    findings_sheet = workbook["Findings"]
    header_row = [cell.value for cell in findings_sheet[1]]
    assert header_row == [
        "Severity",
        "Category",
        "Expected?",
        "Worksheet",
        "Cell",
        "Region",
        "Actual Value",
        "Comparison Basis",
        "Comparator Value",
        "Deviation %",
        "Reason",
        "Suggested Clarification",
    ]

    first_data_row = [cell.value for cell in findings_sheet[2]]
    assert first_data_row[0] == "HIGH"
    assert first_data_row[1] == "Benchmark Comparison"
    assert first_data_row[2] == "No"

    second_data_row = [cell.value for cell in findings_sheet[3]]
    assert second_data_row[2] == "Yes"


def test_write_report_includes_benchmark_coverage_when_present(tmp_path):
    coverage = BenchmarkCoverage(
        total_fields=160,
        matched_fields=142,
        unmatched_by_sheet={"Extras": 12, "Labour": 6},
        unmatched_by_field={"Day Rate": 8, "Call-Out Fee": 5},
    )

    result = SupplierAnalysisResult(
        supplier_name="Acme Ltd",
        findings=[],
        benchmark_coverage=coverage,
    )

    output_path = tmp_path / "acme_report.xlsx"
    ReportWriter().write_report(result, str(output_path))

    workbook = openpyxl.load_workbook(output_path)
    summary_rows = _summary_rows(workbook)

    assert ("Benchmark Coverage", None) in summary_rows
    assert ("Fields Submitted", 160) in summary_rows
    assert ("Matched to a Benchmark Value", 142) in summary_rows
    assert ("Match Rate", "88.8%") in summary_rows
    assert ("Day Rate", 8) in summary_rows


def test_write_report_omits_coverage_section_when_absent(tmp_path):
    result = SupplierAnalysisResult(supplier_name="Acme Ltd", findings=[])

    output_path = tmp_path / "acme_report.xlsx"
    ReportWriter().write_report(result, str(output_path))

    workbook = openpyxl.load_workbook(output_path)
    summary_rows = _summary_rows(workbook)

    assert not any(row[0] == "Benchmark Coverage" for row in summary_rows)


def test_write_report_creates_output_directory_if_missing(tmp_path):
    nested_path = tmp_path / "nested" / "dir" / "report.xlsx"

    result = SupplierAnalysisResult(supplier_name="Acme Ltd", findings=[])
    ReportWriter().write_report(result, str(nested_path))

    assert nested_path.exists()


def test_write_report_neutralises_excel_formula_injection(tmp_path):
    """
    A supplier's submitted cell value (or a maliciously-named
    supplier file) starting with a formula-trigger character must
    never become a live formula in the generated report - openpyxl
    (and Excel itself) treats a leading '=', '+', '-', or '@' as the
    start of a formula unless it's escaped.
    """

    findings = [
        Finding(
            severity=Severity.HIGH,
            worksheet_name="=HYPERLINK(\"http://evil.example\")",
            cell_reference="B2",
            actual_value='=cmd|"/c calc"!A1',
            reason="test",
            category=FindingCategory.TENDER_RESPONSE_CHECK.value,
            comparator_value="+2+5",
        ),
    ]

    result = SupplierAnalysisResult(
        supplier_name="=SUM(1,1)", findings=findings
    )

    output_path = tmp_path / "report.xlsx"
    ReportWriter().write_report(result, str(output_path))

    workbook = openpyxl.load_workbook(output_path)
    findings_sheet = workbook["Findings"]

    worksheet_name_cell = findings_sheet["D2"]
    actual_value_cell = findings_sheet["G2"]
    comparator_value_cell = findings_sheet["I2"]
    supplier_name_cell = workbook["Summary"]["B1"]

    for cell in (
        worksheet_name_cell,
        actual_value_cell,
        comparator_value_cell,
        supplier_name_cell,
    ):
        # data_type 's' = plain text, 'f' = a live formula. Every one
        # of these traces back to attacker-controllable input, so
        # none of them may come back as a formula.
        assert cell.data_type == "s"

    # The apostrophe-escaped text is still recognisably the original
    # value for anyone reading the report, just inert.
    assert actual_value_cell.value == '\'=cmd|"/c calc"!A1'
