# tests/test_workbook_parser.py

import openpyxl

from parsers.workbook_parser import WorkbookParser


class _FakeComputedCell:
    def __init__(self, value):
        self.value = value


class _FakeComputedWorksheet:
    """
    Stands in for the data_only=True worksheet WorkbookParser reads
    computed formula results from - lets the merge logic itself be
    tested directly without depending on openpyxl (which never
    evaluates formulas) or real Excel to produce a cached value.
    """

    def __init__(self, values_by_coordinate):
        self._values_by_coordinate = values_by_coordinate

    def cell(self, row, column):
        return _FakeComputedCell(
            self._values_by_coordinate.get((row, column))
        )


def test_formula_cell_falls_back_to_formula_text_when_no_cached_value(
    tmp_path,
):
    """
    openpyxl itself never computes formulas, so a workbook it writes
    has no cached result for a formula cell - this is the case for
    any workbook that's never actually been opened/saved in Excel.
    Behaviour here must be unchanged from before the fix: the raw
    formula text is kept (still unusable for comparison, but visible
    rather than silently substituted with something wrong).
    """

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Labour"
    worksheet["B2"] = 200
    worksheet["C2"] = "=B2*1.1"

    file_path = tmp_path / "formula.xlsx"
    workbook.save(file_path)

    info = WorkbookParser().load_workbook(str(file_path))
    cell = info.get_worksheet("Labour").get_cell("C2")

    assert cell.has_formula is True
    assert cell.value == "=B2*1.1"


def test_parse_worksheet_uses_computed_value_for_formula_cells(tmp_path):
    """
    When a cached computed result IS available (the normal case for
    any workbook a supplier actually opened and saved in Excel), the
    parser must use that numeric result instead of the formula text -
    otherwise every formula-driven cell silently fails every numeric
    comparison downstream without ever being flagged as unreadable.
    """

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Labour"
    worksheet["B2"] = 200
    worksheet["C2"] = "=B2*1.1"

    parser = WorkbookParser()

    # C2 is row 2, column 3.
    fake_computed_worksheet = _FakeComputedWorksheet({(2, 3): 220.0})

    worksheet_info = parser._parse_worksheet(
        worksheet, fake_computed_worksheet
    )

    cell = worksheet_info.get_cell("C2")
    assert cell.has_formula is True
    assert cell.value == 220.0


def test_non_formula_cells_are_unaffected_by_computed_values_worksheet(
    tmp_path,
):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Labour"
    worksheet["B2"] = 200

    parser = WorkbookParser()

    # Even if the computed-values lookup had something at B2, a
    # non-formula cell's own literal value must always win.
    fake_computed_worksheet = _FakeComputedWorksheet({(2, 2): 999})

    worksheet_info = parser._parse_worksheet(
        worksheet, fake_computed_worksheet
    )

    cell = worksheet_info.get_cell("B2")
    assert cell.has_formula is False
    assert cell.value == 200
